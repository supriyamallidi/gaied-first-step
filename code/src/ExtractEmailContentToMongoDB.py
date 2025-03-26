import io
import multiprocessing
import os
import tempfile
import zipfile
from datetime import datetime
from email import message_from_bytes

import pytesseract
from docx import Document
from openpyxl import load_workbook
from PIL import Image
from pymongo import MongoClient
from PyPDF2 import PdfReader

# Configure Tesseract OCR
pytesseract.pytesseract.tesseract_cmd = r"./resources/tesseract.exe"

# MongoDB Configuration
MONGO_CONNECTION_STRING = "mongodb://localhost:27017/"
DB_NAME = "email_routing"
COLLECTION_NAME = "email_datasets"


def extract_text_from_image(image_bytes):
    """Extracts text from an image using OCR."""
    try:
        image = Image.open(io.BytesIO(image_bytes))
        return pytesseract.image_to_string(image)
    except Exception as e:
        return f"OCR Error: {e}"


def process_msg_attachment(attachment_name, attachment_content):
    """Processes .msg attachment and extracts its content."""
    try:
        # Create a temporary file to process the .msg
        with tempfile.NamedTemporaryFile(delete=False, suffix='.msg') as tmp:
            tmp.write(attachment_content)
            tmp_path = tmp.name

        # Process the .msg file similar to main email processing
        with open(tmp_path, 'rb') as msg_file:
            msg = message_from_bytes(msg_file.read())

        content = {
            'type': 'msg',
            'from': msg.get('From', 'Unknown Sender'),
            'subject': msg.get('Subject', 'No Subject'),
            'body': '',
            'attachments': []
        }

        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/plain":
                    payload = part.get_payload(decode=True)
                    if payload:
                        content['body'] += payload.decode(
                            'utf-8', errors='ignore')
                elif part.get_content_disposition() == 'attachment':
                    nested_attachment_name = part.get_filename()
                    nested_attachment_content = part.get_payload(decode=True)
                    if nested_attachment_name and nested_attachment_content:
                        nested_attachment = {
                            'name': nested_attachment_name,
                            'size': len(nested_attachment_content)
                        }
                        # Process the nested attachment
                        nested_attachment.update(process_attachment(
                            nested_attachment_name, nested_attachment_content))
                        content['attachments'].append(nested_attachment)
        else:
            payload = msg.get_payload(decode=True)
            if payload:
                content['body'] = payload.decode('utf-8', errors='ignore')

        # Clean up temporary file
        os.unlink(tmp_path)

        return content
    except Exception as e:
        return {
            'type': 'msg_error',
            'content': f"Error processing MSG attachment {attachment_name}: {e}"
        }


def process_attachment(attachment_name, attachment_content):
    """Processes attachment content and extracts text based on file type."""
    ext = os.path.splitext(attachment_name)[
        1].lower() if attachment_name else ''
    try:
        if ext in [".txt", ".csv", ".html"]:
            return {
                'type': 'text',
                'content': attachment_content.decode("utf-8", errors="ignore")
            }
        elif ext == ".docx":
            with io.BytesIO(attachment_content) as file_stream:
                doc = Document(file_stream)
                return {
                    'type': 'docx',
                    'content': "\n".join(paragraph.text for paragraph in doc.paragraphs)
                }
        elif ext == ".xlsx":
            with io.BytesIO(attachment_content) as file_stream:
                wb = load_workbook(file_stream)
                text = ""
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    text += f"Sheet: {sheet}\n"
                    for row in ws.iter_rows(values_only=True):
                        text += ", ".join(str(cell)
                                          if cell else "" for cell in row) + "\n"
                return {
                    'type': 'xlsx',
                    'content': text
                }
        elif ext == ".pdf":
            pdf_reader = PdfReader(io.BytesIO(attachment_content))
            return {
                'type': 'pdf',
                'content': "\n".join(page.extract_text() for page in pdf_reader.pages)
            }
        elif ext in [".jpg", ".png", ".jpeg", ".bmp", ".tiff"]:
            return {
                'type': 'image',
                'content': extract_text_from_image(attachment_content)
            }
        elif ext == ".zip":
            with zipfile.ZipFile(io.BytesIO(attachment_content)) as zip_file:
                text = ""
                for zip_info in zip_file.infolist():
                    with zip_file.open(zip_info) as file:
                        try:
                            file_content = file.read().decode("utf-8", errors="ignore")
                            text += f"Zip Entry: {zip_info.filename}\n{file_content}\n"
                        except UnicodeDecodeError:
                            text += f"Binary content in {zip_info.filename}\n"
                return {
                    'type': 'zip',
                    'content': text
                }
        elif ext == ".msg":
            return process_msg_attachment(attachment_name, attachment_content)
        else:
            return {
                'type': 'unsupported',
                'content': f"Unsupported file type: {attachment_name or 'unnamed_attachment'}."
            }
    except Exception as e:
        return {
            'type': 'error',
            'content': f"Error processing attachment {attachment_name or 'unnamed_attachment'}: {e}"
        }


def process_email_file(file_path):
    """Processes a single email file and returns data for MongoDB."""
    try:
        filename = os.path.basename(file_path)
        with open(file_path, "rb") as eml_file:
            email_message = message_from_bytes(eml_file.read())

        # Extract metadata and body
        from_address = email_message["From"] or "Unknown Sender"
        subject = email_message["Subject"] or "No Subject"
        date = email_message["Date"] or datetime.now().isoformat()
        body = ""
        attachments = []
        processed_attachments = []

        if email_message.is_multipart():
            for part in email_message.walk():
                if part.get_content_type() == "text/plain":
                    payload = part.get_payload(decode=True)
                    if payload:
                        body += payload.decode("utf-8", errors="ignore")
                elif part.get_content_disposition() == "attachment":
                    attachments.append(part)
        else:
            payload = email_message.get_payload(decode=True)
            if payload:
                body = payload.decode("utf-8", errors="ignore")

        body = body or "No Body Content"

        # Process attachments
        for attachment in attachments:
            attachment_name = attachment.get_filename()
            attachment_content = attachment.get_payload(decode=True)
            if attachment_content:  # Process even if attachment_name is None
                processed_attachment = {
                    "name": attachment_name or "unnamed_attachment",
                    "size": len(attachment_content)
                }
                processed_attachment.update(process_attachment(
                    attachment_name, attachment_content))
                processed_attachments.append(processed_attachment)

        # Return document data
        return {
            "filename": filename,
            "from": from_address,
            "subject": subject,
            "date": date,
            "body": body,
            "attachments": processed_attachments,
            "processing_date": datetime.now(),
            "status": "processed"
        }

    except Exception as e:
        print(f"Error processing {file_path}: {e}")
        return None


def worker(file_path):
    """Worker function that handles MongoDB connection for each process."""
    try:
        client = MongoClient(MONGO_CONNECTION_STRING)
        db = client[DB_NAME]
        collection = db[COLLECTION_NAME]

        email_data = process_email_file(file_path)
        if email_data:
            # Update existing document or insert new one
            result = collection.update_one(
                {"filename": email_data["filename"]},
                {"$set": email_data},
                upsert=True
            )
            if result.upserted_id:
                print(
                    f"Inserted new document for {file_path} (ID: {result.upserted_id})")
            else:
                print(f"Updated existing document for {file_path}")
    except Exception as e:
        print(f"Error processing {file_path}: {e}")
    finally:
        client.close()


def process_files_in_parallel(msg_folder):
    """Processes email files in parallel using multiprocessing."""
    file_paths = [os.path.join(msg_folder, filename) for filename in os.listdir(msg_folder)
                  if filename.endswith((".eml", ".msg"))]

    with multiprocessing.Pool(processes=multiprocessing.cpu_count()) as pool:
        pool.map(worker, file_paths)


def ensure_db_and_collection(uri, db_name, collection_name):
    client = MongoClient(uri)
    db = client[db_name]
    if collection_name not in db.list_collection_names():
        db.create_collection(collection_name)
    client.close()


if __name__ == "__main__":
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__)))
    RESOURCES_DIR = os.path.join(BASE_DIR, "resources")
    msg_folder_path = os.path.join(RESOURCES_DIR, 'emails')

    print(f"Message Folder Path: {msg_folder_path}")
    print("Starting email processing...")
    ensure_db_and_collection(MONGO_CONNECTION_STRING, DB_NAME, COLLECTION_NAME)
    process_files_in_parallel(msg_folder_path)
